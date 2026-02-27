#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JobSearch AI — Streamlit webbapp
Automatiserad jobbsökning via JobTech + AI-matchning via Gemini / g4f
"""

import streamlit as st
import requests
import pandas as pd
import time
import re
import io
import concurrent.futures
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Lazy import g4f (may not be needed if user has Gemini key)
def _get_g4f():
    import g4f
    return g4f

# ============================================================
# CONSTANTS
# ============================================================
JOBTECH_URL = "https://jobsearch.api.jobtechdev.se/search"

KOMMUN_IDS = {
    "Stockholm": "0180",
    "Göteborg": "1480",
    "Malmö": "1280",
    "Linköping": "0580",
    "Västerås": "1980",
    "Örebro": "1880",
    "Umeå": "2480",
    "Remote": "0000",
}

MODEL_CHAIN = ["openai-large", "gemini", "openai", "mistral", "deepseek"]

COL_WIDTHS = [10, 45, 30, 20, 12, 60, 50, 30]
COL_NAMES = ["Poäng", "Titel", "Företag", "Ort", "Datum", "Motivering", "URL", "ID"]

BATCH_SIZE = 5
PAUSE_BETWEEN_BATCHES = 15
RATE_LIMIT_WAIT = 65

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="JobSearch AI",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# DARK THEME CSS
# ============================================================
st.markdown("""
<style>
    .stApp { background-color: #1E1E1E; }
    .block-container { padding-top: 2rem; }
    h1, h2, h3, p, span, label, li { color: #FFFFFF !important; }
    .stTextInput>div>div>input, .stTextArea>div>div>textarea {
        background-color: #333333 !important;
        color: #FFFFFF !important;
        border: 1px solid #555555 !important;
    }
    .stButton>button {
        background-color: #0078D4 !important;
        color: white !important;
        font-size: 1.2rem !important;
        font-weight: bold !important;
        padding: 0.6rem 2rem !important;
        border: none !important;
        border-radius: 6px !important;
        width: 100% !important;
    }
    .stButton>button:hover {
        background-color: #005FA3 !important;
    }
    .stDownloadButton>button {
        background-color: #4EC94E !important;
        color: white !important;
        font-size: 1.1rem !important;
        font-weight: bold !important;
        border: none !important;
        border-radius: 6px !important;
        width: 100% !important;
    }
    .log-box {
        background-color: #2D2D2D;
        color: #4EC94E;
        font-family: 'Consolas', monospace;
        font-size: 0.85rem;
        padding: 12px;
        border-radius: 6px;
        max-height: 300px;
        overflow-y: auto;
        white-space: pre-wrap;
    }
    .metric-card {
        background-color: #2D2D2D;
        padding: 16px;
        border-radius: 8px;
        text-align: center;
    }
    .metric-card h2 { color: #FFD700 !important; margin: 0; }
    .metric-card p { color: #FFFFFF !important; margin: 4px 0 0 0; font-size: 0.9rem; }
    div[data-testid="stSidebar"] { background-color: #2D2D2D !important; }
    div[data-testid="stSidebar"] label { color: #FFFFFF !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# SESSION STATE
# ============================================================
if "log" not in st.session_state:
    st.session_state.log = []
if "results" not in st.session_state:
    st.session_state.results = None
if "running" not in st.session_state:
    st.session_state.running = False
if "working_model" not in st.session_state:
    st.session_state.working_model = None

def add_log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    st.session_state.log.append(f"[{ts}] {msg}")

# ============================================================
# SIDEBAR — SETTINGS
# ============================================================
with st.sidebar:
    st.markdown("## ⚙️ Inställningar")

    cities_input = st.text_input(
        "Städer (komma-sep)",
        value="Stockholm, Göteborg, Malmö, Remote",
    )
    roles_input = st.text_input(
        "Roller / sökord",
        value="Data Analyst, AI Engineer, Python Developer",
    )
    exclude_input = st.text_input(
        "Exkludera ord",
        value="senior, chef, praktik, säljare",
    )
    max_jobs = st.number_input(
        "Max jobb att analysera", min_value=1, max_value=500, value=50, step=10
    )

    st.markdown("---")
    st.markdown("### 🔑 Gemini API")
    gemini_key = st.text_input(
        "API-nyckel (valfritt)",
        type="password",
        help="Med egen nyckel: 10× snabbare (parallellt). Utan: gratis via g4f (långsammare).",
    )
    st.markdown(
        '[Hämta gratis nyckel →](https://aistudio.google.com/app/api-keys)',
        unsafe_allow_html=True,
    )

    st.markdown("---")
    cv_text = st.text_area(
        "CV / Profil",
        height=200,
        value=(
            "Civilingenjör med 5 års erfarenhet inom mjukvaruutveckling och dataanalys. "
            "Kompetenser: Python, SQL, Machine Learning, REST API:er, Docker, Git. "
            "Tidigare roller: systemutvecklare och tech lead i agila team. "
            "Utbildning: MSc Datateknik, Chalmers. "
            "Skriv in ditt eget CV här för bästa matchning."
        ),
    )

# ============================================================
# MAIN AREA
# ============================================================
st.markdown("# 🚀 JobSearch AI")
st.markdown("Automatiserad jobbsökning via **JobTech API** + AI-matchning")

# ============================================================
# JOBTECH FETCH
# ============================================================
def fetch_jobs(cities, roles, excludes, progress_bar, status_text):
    seen_ids = set()
    all_jobs = []
    city_list = [c.strip() for c in cities.split(",") if c.strip()]
    role_list = [r.strip() for r in roles.split(",") if r.strip()]
    exclude_list = [e.strip().lower() for e in excludes.split(",") if e.strip()]

    total_combos = len(city_list) * len(role_list)
    combo_idx = 0

    for city in city_list:
        city_key = city.strip().capitalize()
        kommun_id = KOMMUN_IDS.get(city_key)

        for role in role_list:
            combo_idx += 1
            status_text.text(f"🔍 Söker '{role}' i {city_key}...")
            progress_bar.progress(combo_idx / total_combos)
            add_log(f"🔍 Söker '{role}' i {city_key}...")

            params = {"q": role, "limit": 100}
            if kommun_id and kommun_id != "0000":
                params["municipality"] = kommun_id
            elif kommun_id == "0000":
                params["remote"] = "true"

            try:
                resp = requests.get(
                    JOBTECH_URL, params=params, timeout=20,
                    headers={"accept": "application/json"}
                )
                resp.raise_for_status()
                data = resp.json()
            except Exception as e:
                add_log(f"  ⚠️ API-fel: {e}")
                time.sleep(0.3)
                continue

            hits = data.get("hits", [])
            count = 0
            for h in hits:
                jid = h.get("id", "")
                headline = h.get("headline", "")
                if jid in seen_ids:
                    continue
                if any(ex in headline.lower() for ex in exclude_list):
                    continue
                seen_ids.add(jid)
                count += 1
                all_jobs.append({
                    "id": jid,
                    "title": headline,
                    "company": (h.get("employer") or {}).get("name", "Okänt"),
                    "city": (h.get("workplace_address") or {}).get("municipality", city_key),
                    "date": (h.get("publication_date") or "")[:10],
                    "url": h.get("webpage_url", ""),
                    "desc": (h.get("description") or {}).get("text", "")[:600],
                })

            add_log(f"  → {count} nya (totalt {len(all_jobs)})")
            time.sleep(0.3)

    add_log(f"📋 Totalt {len(all_jobs)} unika jobb hämtade.")
    return all_jobs

# ============================================================
# AI ANALYSIS
# ============================================================
def build_prompt(job, cv):
    return (
        f"Du är en jobbmatchningsexpert. Analysera hur väl detta jobb matchar kandidatens CV.\n\n"
        f"JOBB:\nTitel: {job['title']}\nFöretag: {job['company']}\nOrt: {job['city']}\n"
        f"Beskrivning: {job['desc']}\n\n"
        f"CV:\n{cv}\n\n"
        f"Svara EXAKT i detta format (inget annat):\n"
        f"POÄNG: XX.XX\n"
        f"MOTIVERING: (kort motivering på svenska, max 2 meningar)\n"
    )

def parse_ai_response(text):
    score_match = re.search(r"PO[ÄA]NG:\s*([\d]{1,3}(?:[.,]\d{1,2})?)", text)
    reason_match = re.search(r"MOTIVERING:\s*(.+)", text, re.DOTALL)
    if score_match:
        score = float(score_match.group(1).replace(",", "."))
        score = min(100.0, max(0.0, score))
        reason = reason_match.group(1).strip()[:300] if reason_match else "Ingen motivering."
        return round(score, 2), reason
    return None

def try_gemini_api(prompt, api_key):
    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"gemini-flash-latest:generateContent?key={api_key}"
    )
    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    try:
        resp = requests.post(url, json=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        text = data["candidates"][0]["content"]["parts"][0]["text"]
        return parse_ai_response(text)
    except Exception:
        return None

def try_g4f(prompt):
    g4f = _get_g4f()
    models = ([st.session_state.working_model] if st.session_state.working_model else []) + MODEL_CHAIN
    for model in models:
        try:
            response = g4f.ChatCompletion.create(
                model=model,
                provider=g4f.Provider.PollinationsAI,
                messages=[{"role": "user", "content": prompt}],
            )
            text = response if isinstance(response, str) else "".join(response)
            result = parse_ai_response(text)
            if result:
                st.session_state.working_model = model
                return result
        except Exception:
            continue
    return None

def analyze_one(job, cv, gemini_key):
    prompt = build_prompt(job, cv)

    # Try Gemini first if key provided
    if gemini_key:
        result = try_gemini_api(prompt, gemini_key)
        if result:
            return result

    # Fallback to g4f
    result = try_g4f(prompt)
    if result:
        return result

    return 0.0, "AI-analys misslyckades."

def analyze_parallel(jobs, cv, gemini_key, progress_bar, status_text):
    """Parallel analysis using Gemini API — 10 threads."""
    results = [None] * len(jobs)
    total = len(jobs)
    completed = [0]
    lock = threading.Lock()

    def _process(idx, job):
        score, reason = analyze_one(job, cv, gemini_key)
        results[idx] = {
            "Poäng": score, "Titel": job["title"], "Företag": job["company"],
            "Ort": job["city"], "Datum": job["date"], "Motivering": reason,
            "URL": job["url"], "ID": job["id"],
        }
        with lock:
            completed[0] += 1

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as pool:
        futures = [pool.submit(_process, i, job) for i, job in enumerate(jobs)]
        while not all(f.done() for f in futures):
            n = completed[0]
            progress_bar.progress(min(n / total, 1.0))
            status_text.text(f"⚡ Gemini: {n}/{total} jobb klara")
            time.sleep(0.3)
        concurrent.futures.wait(futures)

    progress_bar.progress(1.0)
    status_text.text(f"⚡ Gemini: {total}/{total} klara!")
    add_log(f"⚡ Alla {total} jobb analyserade parallellt.")
    return [r for r in results if r is not None]

def analyze_sequential(jobs, cv, gemini_key, progress_bar, status_text):
    """Sequential analysis via g4f with pauses."""
    results = []
    total = len(jobs)

    for i, job in enumerate(jobs):
        status_text.text(f"🤖 Analyserar {i+1}/{total}: {job['title'][:50]}...")
        progress_bar.progress((i + 1) / total)

        score, reason = analyze_one(job, cv, gemini_key)
        results.append({
            "Poäng": score, "Titel": job["title"], "Företag": job["company"],
            "Ort": job["city"], "Datum": job["date"], "Motivering": reason,
            "URL": job["url"], "ID": job["id"],
        })

        # Pause between batches for g4f
        if (i + 1) % BATCH_SIZE == 0 and i + 1 < total:
            add_log(f"  ⏳ Pausar {PAUSE_BETWEEN_BATCHES}s...")
            status_text.text(f"⏳ Pausar {PAUSE_BETWEEN_BATCHES}s (rate-limit)...")
            time.sleep(PAUSE_BETWEEN_BATCHES)

    add_log(f"🤖 Alla {total} jobb analyserade sekventiellt.")
    return results

# ============================================================
# EXCEL BUILDER
# ============================================================
def build_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobbmatchning"

    header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    cell_font = Font(name="Segoe UI", size=10, color="FFFFFF")
    dark_fill_1 = PatternFill(start_color="1E1E1E", end_color="1E1E1E", fill_type="solid")
    dark_fill_2 = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="444444"), right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"), bottom=Side(style="thin", color="444444"),
    )
    align = Alignment(vertical="center", wrap_text=True)

    for col_idx, name in enumerate(COL_NAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = thin_border

    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(COL_NAMES, 1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = align
            cell.border = thin_border
            cell.fill = dark_fill_2 if row_idx % 2 == 0 else dark_fill_1

    for col_idx, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ============================================================
# RUN BUTTON
# ============================================================
run_col1, run_col2, run_col3 = st.columns([1, 2, 1])
with run_col2:
    run_clicked = st.button("🚀  KÖR SÖKNING!", use_container_width=True)

if run_clicked and not st.session_state.running:
    if not cities_input or not roles_input or not cv_text.strip():
        st.error("❌ Fyll i städer, roller och CV!")
    else:
        st.session_state.running = True
        st.session_state.log = []
        st.session_state.results = None

        # --- Phase 1: Fetch ---
        st.markdown("### 📡 Hämtar jobb...")
        prog1 = st.progress(0)
        stat1 = st.empty()

        jobs = fetch_jobs(cities_input, roles_input, exclude_input, prog1, stat1)

        if not jobs:
            st.warning("⚠️ Inga jobb hittades.")
            st.session_state.running = False
        else:
            # Limit
            if len(jobs) > max_jobs:
                add_log(f"✂️ Begränsar till {max_jobs} jobb (av {len(jobs)}).")
                jobs = jobs[:max_jobs]

            # --- Phase 2: AI ---
            use_gemini = bool(gemini_key)
            mode = "⚡ Gemini (parallellt)" if use_gemini else "🤖 g4f (sekventiellt)"
            st.markdown(f"### 🤖 AI-analys — {mode}")
            prog2 = st.progress(0)
            stat2 = st.empty()

            if use_gemini:
                results = analyze_parallel(jobs, cv_text, gemini_key, prog2, stat2)
            else:
                results = analyze_sequential(jobs, cv_text, gemini_key, prog2, stat2)

            if results:
                df = pd.DataFrame(results)
                df.sort_values("Poäng", ascending=False, inplace=True)
                df.reset_index(drop=True, inplace=True)
                st.session_state.results = df
                add_log(f"✅ Klart! {len(results)} jobb analyserade.")

            st.session_state.running = False
            st.rerun()

# ============================================================
# RESULTS DISPLAY
# ============================================================
if st.session_state.results is not None:
    df = st.session_state.results

    st.markdown("---")

    # Metrics
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f'<div class="metric-card"><h2>{len(df)}</h2><p>Jobb analyserade</p></div>', unsafe_allow_html=True)
    with col2:
        top_score = df["Poäng"].max() if len(df) > 0 else 0
        st.markdown(f'<div class="metric-card"><h2>{top_score}</h2><p>Högsta poäng</p></div>', unsafe_allow_html=True)
    with col3:
        avg_score = df["Poäng"].mean() if len(df) > 0 else 0
        st.markdown(f'<div class="metric-card"><h2>{avg_score:.1f}</h2><p>Snittpoäng</p></div>', unsafe_allow_html=True)
    with col4:
        good = len(df[df["Poäng"] >= 70])
        st.markdown(f'<div class="metric-card"><h2>{good}</h2><p>Poäng ≥ 70</p></div>', unsafe_allow_html=True)

    st.markdown("### 📊 Resultat")

    # Display table
    display_df = df[["Poäng", "Titel", "Företag", "Ort", "Datum", "Motivering"]].copy()
    st.dataframe(display_df, use_container_width=True, height=400)

    # Top matches with links
    st.markdown("### 🏆 Topp-matchningar")
    for _, row in df.head(10).iterrows():
        score = row["Poäng"]
        color = "#4EC94E" if score >= 80 else "#FFD700" if score >= 60 else "#FF6B6B"
        st.markdown(
            f'<span style="color:{color};font-weight:bold">{score}</span> — '
            f'**{row["Titel"]}** @ {row["Företag"]} ({row["Ort"]})  \n'
            f'<small>{row["Motivering"]}</small>  \n'
            f'[🔗 Visa annons]({row["URL"]})',
            unsafe_allow_html=True,
        )
        st.markdown("")

    # Excel download
    st.markdown("### 💾 Ladda ner")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    excel_buf = build_excel(df)
    st.download_button(
        label="📥 Ladda ner Excel",
        data=excel_buf,
        file_name=f"jobb_{ts}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ============================================================
# LOG
# ============================================================
if st.session_state.log:
    with st.expander("📋 Logg", expanded=False):
        log_text = "\n".join(st.session_state.log)
        st.markdown(f'<div class="log-box">{log_text}</div>', unsafe_allow_html=True)

# ============================================================
# FOOTER
# ============================================================
st.markdown("---")
st.markdown(
    '<p style="text-align:center;color:#888;font-size:0.8rem;">'
    'JobSearch AI — Powered by JobTech API + Gemini/g4f'
    '</p>',
    unsafe_allow_html=True,
)
